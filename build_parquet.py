#!/usr/bin/env python3
"""Build embedded Parquet from ML-TTS Audio Samples xlsx + audio files."""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import gdown
import openpyxl
import pyarrow as pa
import pyarrow.parquet as pq
from datasets import Dataset, Features, Value


DRIVE_ID_RE = re.compile(r"/file/d/([a-zA-Z0-9_-]+)")
AUDIO_EXTENSIONS = {".wav", ".mp3", ".m4a", ".flac", ".ogg", ".webm", ".aac"}


def extract_drive_id(url: str) -> str:
    match = DRIVE_ID_RE.search(url)
    if not match:
        raise ValueError(f"Could not parse Google Drive file id from: {url}")
    return match.group(1)


def parse_sheet_name(sheet_name: str) -> tuple[str, str]:
    speaker = sheet_name.split("(")[0].strip()
    lang_code = "en" if "(E)" in sheet_name else "hi"
    return speaker, lang_code


def parse_xlsx(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    rows: list[dict] = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "Home":
            continue

        speaker, language = parse_sheet_name(sheet_name)
        ws = wb[sheet_name]

        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None or not isinstance(row[0], (int, float)):
                continue

            script_id = int(row[0])
            drive_url = row[6]
            if not drive_url:
                raise ValueError(f"Missing Drive link for {speaker} script {script_id}")

            rows.append(
                {
                    "script_id": script_id,
                    "language": language,
                    "speaker": speaker,
                    "domain": str(row[1] or ""),
                    "named_entity": str(row[2] or ""),
                    "text": str(row[3] or "").strip(),
                    "transliteration": str(row[4] or ""),
                    "meaning": str(row[5] or "").strip(),
                    "source_url": str(drive_url),
                    "drive_id": extract_drive_id(str(drive_url)),
                }
            )

    wb.close()
    rows.sort(key=lambda r: (r["language"], r["script_id"], r["speaker"]))
    return rows


def detect_extension(path: Path) -> str:
    header = path.read_bytes()[:12]
    if header[:4] == b"RIFF":
        return ".wav"
    if header[:3] == b"ID3" or header[:2] in (b"\xff\xfb", b"\xff\xf3", b"\xff\xf2"):
        return ".mp3"
    if len(header) >= 8 and header[4:8] == b"ftyp":
        return ".m4a"
    if header[:4] == b"fLaC":
        return ".flac"
    if header[:4] == b"OggS":
        return ".ogg"
    return path.suffix or ".bin"


def index_local_audio(audio_root: Path) -> dict[tuple[str, int], Path]:
    """Map (speaker_lower, script_id) -> audio file path."""
    index: dict[tuple[str, int], Path] = {}
    if not audio_root.exists():
        return index

    patterns = [
        re.compile(r"^(?P<speaker>.+?)[-_](?P<id>\d+)\.", re.IGNORECASE),
        re.compile(r"^(?P<speaker>.+?)[-_](?P<id>\d+)$", re.IGNORECASE),
    ]

    for path in audio_root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in AUDIO_EXTENSIONS:
            continue

        stem = path.stem
        for pattern in patterns:
            match = pattern.match(stem)
            if match:
                speaker = match.group("speaker").strip().lower()
                script_id = int(match.group("id"))
                index[(speaker, script_id)] = path
                break

    return index


def find_local_audio(row: dict, audio_root: Path, local_index: dict[tuple[str, int], Path]) -> Path | None:
    speaker_key = row["speaker"].lower()
    direct = local_index.get((speaker_key, row["script_id"]))
    if direct:
        return direct

    # Fallback: canonical cache filename from prior downloads
    for path in audio_root.glob(
        f"{row['language']}_{row['script_id']:02d}_{row['speaker']}.*"
    ):
        if path.suffix.lower() in AUDIO_EXTENSIONS:
            return path

    return None


def download_from_drive(row: dict, audio_dir: Path) -> Path:
    import requests

    audio_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{row['language']}_{row['script_id']:02d}_{row['speaker']}"
    out_base = audio_dir / filename

    existing = [p for p in audio_dir.glob(f"{filename}.*") if p.suffix.lower() in AUDIO_EXTENSIONS]
    if existing:
        return existing[0]

    tmp_path = audio_dir / f"{filename}.bin"
    file_id = row["drive_id"]

    def try_requests_download() -> bool:
        session = requests.Session()
        url = "https://docs.google.com/uc?export=download"
        response = session.get(url, params={"id": file_id}, timeout=10)
        token = None
        for key, value in response.cookies.items():
            if key.startswith("download_warning"):
                token = value
                break
        if token:
            response = session.get(url, params={"id": file_id, "confirm": token}, timeout=30)
        content = response.content
        if not content or content[:15].lower().startswith(b"<!doctype") or b"<html" in content[:500].lower():
            return False
        tmp_path.write_bytes(content)
        return tmp_path.stat().st_size > 0

    last_error: Exception | None = None
    if try_requests_download():
        ext = detect_extension(tmp_path)
        final_path = out_base.with_suffix(ext)
        tmp_path.rename(final_path)
        return final_path

    for url in (row["source_url"], f"https://drive.google.com/uc?id={file_id}"):
        try:
            gdown.download(url, str(tmp_path), quiet=True, fuzzy=True, use_cookies=True)
            if tmp_path.exists() and tmp_path.stat().st_size > 0:
                header = tmp_path.read_bytes()[:500]
                if b"<html" in header.lower():
                    tmp_path.unlink()
                    continue
                ext = detect_extension(tmp_path)
                final_path = out_base.with_suffix(ext)
                tmp_path.rename(final_path)
                return final_path
        except Exception as exc:  # noqa: BLE001 - collect and try next URL
            last_error = exc
            if tmp_path.exists():
                tmp_path.unlink()

    raise RuntimeError(
        f"Could not download audio for {row['speaker']} script {row['script_id']}. "
        f"Last error: {last_error}. "
        "Ensure Drive files are shared as 'Anyone with the link'."
    )


def resolve_audio(row: dict, audio_dir: Path, local_index: dict[tuple[str, int], Path], skip_download: bool) -> Path:
    local = find_local_audio(row, audio_dir, local_index)
    if local:
        return local
    if skip_download:
        raise FileNotFoundError(
            f"No local audio for {row['speaker']} script {row['script_id']}. "
            f"Expected something like {row['speaker']}-{row['script_id']:03d}.wav under {audio_dir}"
        )
    return download_from_drive(row, audio_dir)


def build_embedded_records(
    rows: list[dict],
    audio_dir: Path,
    skip_download: bool,
) -> list[dict]:
    local_index = index_local_audio(audio_dir)
    records = []
    missing: list[str] = []

    for i, row in enumerate(rows):
        try:
            audio_path = resolve_audio(row, audio_dir, local_index, skip_download)
        except (FileNotFoundError, RuntimeError) as exc:
            missing.append(f"{row['speaker']} #{row['script_id']} ({row['language']}): {exc}")
            continue

        print(f"[{i + 1}/{len(rows)}] {audio_path.name}")
        records.append(
            {
                "script_id": row["script_id"],
                "language": row["language"],
                "speaker": row["speaker"],
                "domain": row["domain"],
                "named_entity": row["named_entity"],
                "text": row["text"],
                "transliteration": row["transliteration"],
                "meaning": row["meaning"],
                "source_url": row["source_url"],
                "audio_bytes": audio_path.read_bytes(),
                "audio_path": audio_path.name,
            }
        )

    if missing:
        print("\nMissing audio files:", file=sys.stderr)
        for item in missing:
            print(f"  - {item}", file=sys.stderr)
        if not records:
            raise SystemExit(
                "\nNo audio files found. Either:\n"
                "  1. Make the Google Drive folder public ('Anyone with the link'), then re-run\n"
                "  2. Download the folder manually to audio/ using names like Thrisha-001.wav\n"
                "     Folder: https://drive.google.com/drive/folders/1E-3-uCb_XpLq6nY7hhSIUvN1kMR6IWbR?usp=sharing\n"
                "  3. Re-run with --skip-download once files are in place\n"
            )
        print(f"\nWarning: building parquet with {len(records)}/{len(rows)} samples", file=sys.stderr)

    return records


def write_embedded_parquet(records: list[dict], output: Path) -> None:
    audio_struct = pa.StructArray.from_arrays(
        [
            pa.array([r["audio_bytes"] for r in records], type=pa.binary()),
            pa.array([r["audio_path"] for r in records], type=pa.string()),
        ],
        names=["bytes", "path"],
    )

    table = pa.table(
        {
            "script_id": pa.array([r["script_id"] for r in records], type=pa.int32()),
            "language": [r["language"] for r in records],
            "speaker": [r["speaker"] for r in records],
            "domain": [r["domain"] for r in records],
            "named_entity": [r["named_entity"] for r in records],
            "text": [r["text"] for r in records],
            "transliteration": [r["transliteration"] for r in records],
            "meaning": [r["meaning"] for r in records],
            "source_url": [r["source_url"] for r in records],
            "audio": audio_struct,
        }
    )
    pq.write_table(table, output)


def build_metadata_dataset(rows: list[dict]) -> Dataset:
    features = Features(
        {
            "script_id": Value("int32"),
            "language": Value("string"),
            "speaker": Value("string"),
            "domain": Value("string"),
            "named_entity": Value("string"),
            "text": Value("string"),
            "transliteration": Value("string"),
            "meaning": Value("string"),
            "source_url": Value("string"),
            "drive_id": Value("string"),
        }
    )
    return Dataset.from_list(rows, features=features)


def write_manifest(rows: list[dict], manifest_path: Path) -> None:
    manifest_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=Path("ML-TTS Audio Samples.xlsx"))
    parser.add_argument("--audio-dir", type=Path, default=Path("audio"))
    parser.add_argument("--output", type=Path, default=Path("ml_tts_samples_embedded.parquet"))
    parser.add_argument("--manifest", type=Path, default=Path("samples_manifest.json"))
    parser.add_argument(
        "--skip-download",
        action="store_true",
        help="Only use audio files already present under --audio-dir",
    )
    parser.add_argument(
        "--metadata-only",
        action="store_true",
        help="Write text metadata parquet without embedded audio",
    )
    parser.add_argument(
        "--list-expected",
        action="store_true",
        help="Print expected local audio filenames and exit",
    )
    args = parser.parse_args()

    if not args.xlsx.exists():
        print(f"Error: xlsx not found: {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    print(f"Parsing {args.xlsx} ...")
    rows = parse_xlsx(args.xlsx)
    en_count = sum(1 for r in rows if r["language"] == "en")
    hi_count = sum(1 for r in rows if r["language"] == "hi")
    print(f"Found {len(rows)} samples ({en_count} en, {hi_count} hi)")

    if args.list_expected:
        for row in rows:
            print(f"{row['speaker']}-{row['script_id']:03d}  ({row['language']})  {row['text'][:60]}...")
        return

    write_manifest(rows, args.manifest)
    print(f"Wrote manifest -> {args.manifest}")

    if args.metadata_only:
        args.output = args.output.with_name(
            args.output.stem.replace("_embedded", "") + "_metadata.parquet"
            if "_embedded" in args.output.stem
            else "ml_tts_samples_metadata.parquet"
        )
        ds = build_metadata_dataset(rows)
        print(f"Writing parquet -> {args.output}")
        ds.to_parquet(str(args.output))
        row_count = len(ds)
    else:
        print(f"Resolving audio from {args.audio_dir} ...")
        records = build_embedded_records(rows, args.audio_dir, args.skip_download)
        print(f"Writing embedded parquet -> {args.output}")
        write_embedded_parquet(records, args.output)
        row_count = len(records)

    size_mb = args.output.stat().st_size / (1024 * 1024)
    print(f"Done. {row_count} rows, {size_mb:.1f} MB")


if __name__ == "__main__":
    main()
